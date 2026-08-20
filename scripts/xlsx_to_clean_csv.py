#!/usr/bin/env python3
"""
Converte il report .xlsx prodotto da public-claude-semrush-keyword-cleaner
(scripts/semrush_cleaner.py --mode clean, foglio "Tutti i Dati") nel CSV piatto
atteso in input da public-claude-clustering-agent, aggiungendo le colonne Brand e Country
che il cleaner non scrive (servono al downstream: build_sheet.py e le regole di
clustering per lingua).

Uso:
    python scripts/xlsx_to_clean_csv.py \
        --xlsx runs/<slug>/clean/report.xlsx \
        --output runs/<slug>/clean/all_clean.csv \
        --brand Yamamay
"""

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Tutti i Dati"
HEADERS_IN = ["Mercato", "Mercato Code", "Data", "Keyword", "Position", "Search Volume", "URL", "Brand/Not Brand"]
# Versioni recenti del cleaner anticipano anche una colonna "Brand" (dominio), che qui
# viene ignorata e sovrascritta col nome brand passato a --brand.
HEADERS_IN_WITH_DOMAIN_BRAND = ["Brand"] + HEADERS_IN

# Bucket lingua usati dalle regole di public-claude-clustering-agent (rules/it|en|es|fr|de.json)
COUNTRY_BUCKET = {"it": "IT", "es": "ES", "fr": "FR", "de": "DE"}


def main():
    parser = argparse.ArgumentParser(description="XLSX del keyword-cleaner -> CSV piatto per il clustering")
    parser.add_argument("--xlsx", required=True, help="report .xlsx prodotto da semrush_cleaner.py --mode clean")
    parser.add_argument("--output", required=True, help="CSV di output (input per public-claude-clustering-agent)")
    parser.add_argument("--brand", required=True, help="Nome brand da scrivere nella colonna Brand")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"File non trovato: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Foglio '{SHEET_NAME}' non trovato in {xlsx_path} (fogli presenti: {wb.sheetnames})", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    header_list = list(header) if header is not None else None
    if header_list == HEADERS_IN_WITH_DOMAIN_BRAND:
        skip_domain_brand_col = True
    elif header_list == HEADERS_IN:
        skip_domain_brand_col = False
    else:
        print(f"Header inatteso nel foglio '{SHEET_NAME}': {header}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    out_header = HEADERS_IN + ["Brand", "Country"]
    n_rows = 0
    with open(output_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(out_header)
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            values = list(row)
            if skip_domain_brand_col:
                values = values[1:]
            mercato_code = str(values[1] or "").lower()
            country = COUNTRY_BUCKET.get(mercato_code, "EN")
            writer.writerow(values + [args.brand, country])
            n_rows += 1

    print(f"Righe scritte: {n_rows}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
