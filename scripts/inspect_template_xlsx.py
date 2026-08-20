#!/usr/bin/env python3
"""
Diagnostica READ-ONLY del template Sheet, offline: sostituisce la parte "Sheet" di
inspect_template.py, che leggeva il template remoto via Sheets API.

Il template va prima scaricato/esportato in locale con il tool MCP
mcp__claude_ai_Google_Drive__download_file_content (fileId=GOOGLE_SHEET_TEMPLATE_ID,
exportMimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet), da Claude
nella conversazione — questo script non fa alcuna chiamata di rete, solo lettura locale.

Non modifica nulla: solo lettura e stampa.

Uso:
    python scripts/inspect_template_xlsx.py --xlsx runs/<slug>/template/sheet_template.xlsx
"""

import argparse

from openpyxl import load_workbook


def chart_title_text(chart) -> str:
    try:
        runs = chart.title.tx.rich.p[0].r
        return "".join(run.t or "" for run in runs)
    except (AttributeError, IndexError, TypeError):
        return "(senza titolo)"


def main():
    parser = argparse.ArgumentParser(description="Ispeziona (sola lettura) un template Sheet esportato in xlsx")
    parser.add_argument("--xlsx", required=True, help="Percorso locale dell'xlsx esportato dal template Google Sheet")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=False)
    print(f"File: {args.xlsx}")
    print(f"Tab presenti: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n- Tab '{name}' ({ws.max_row}x{ws.max_column})")
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            print(f"    {list(row)}")
        if ws._charts:
            for chart in ws._charts:
                print(f"    chart: {chart_title_text(chart)!r}")


if __name__ == "__main__":
    main()
