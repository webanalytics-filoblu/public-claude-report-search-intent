#!/usr/bin/env python3
"""
Genera l'xlsx completo del report, interamente offline (nessuna chiamata Google): sostituisce
build_sheet.py, che scriveva direttamente su un Google Sheet remoto via Sheets API batchUpdate.

Parte dal **template scaricato da Drive** (--template-xlsx: il Google Sheet template
esportato in locale da Claude con mcp__claude_ai_Google_Drive__download_file_content,
exportMimeType=.../spreadsheetml.sheet — stesso file ispezionato con
inspect_template_xlsx.py), non da un Workbook vuoto: preserva così ogni tab/foglio di
riferimento non gestito da questo script (istruzioni, stile) presente nel template. I tab
gestiti da questo script (brand-secco, clusters, cluster-overview, i tab-cluster/breakdown,
ed eventuali tab di esempio residui — LEGACY_EXAMPLE_TABS) vengono ricreati da zero: se
esistono già nel template con quel nome, vengono rimossi e riscritti.

Claude carica poi il file risultante su Drive con il tool MCP
mcp__claude_ai_Google_Drive__create_file (content_mime_type xlsx, conversione automatica in
Google Sheet) — nessuno script Python fa più chiamate Drive/Sheets dirette.

Differenze deliberate rispetto a build_sheet.py (perdita di funzionalità accettata):
  - Niente più PivotTable native di Sheets (addPivotTable): le aggregazioni per Sotto Cluster
    / Cluster Effettivo / colonna extra sono pre-calcolate qui con pandas.groupby e scritte
    come tabella statica. Se l'utente modifica a mano i dati dopo l'upload, la tabella e il
    grafico NON si ricalcolano da soli (andrebbe rigenerato l'xlsx e ricaricato).
  - Niente più grafici Sheets "linked" apribili da build_slides_pptx.py: qui il grafico è un
    openpyxl.chart.PieChart embedded nello stesso file, statico.

Stessa struttura di tab di prima:
  - "brand-secco": 1 riga di volumi (ultimi 3 anni) + 1 riga di formule Excel per le
    variazioni % YoY (=(A2-B2)/B2, =(B2-C2)/C2).
  - "clusters": risultato completo della clusterizzazione (8 colonne fisse + colonne
    attributo dinamiche, es. Stagionalità/Genere/Brand correlati — derivate da
    attribute_columns_after_sotto_cluster, non una lista fissa) + colonna ausiliaria
    "Cluster Effettivo"/"Cluster Overview" scritta da apply_cluster_plan.
  - "cluster-overview": tabella statica (Cluster Overview -> SUM Search Volume) + torta,
    filtrata sulle sole etichette overview del piano.
  - un tab per ogni entry di --cluster-plan: tabella statica (Sotto Cluster -> SUM Search
    Volume) + torta, filtrata su quella label di "Cluster Effettivo".
  - un tab "Extra <Colonna>" per ogni colonna attributo dinamica non interamente vuota:
    stessa tabella+torta ma senza filtro (source = intero tab "clusters").

Scrive charts_meta.json con schema ridotto rispetto a prima (niente più spreadsheetId/
sheetId/chartId Google, che qui non esistono ancora): { overview: {sheetTitle, totalVolume,
subClusters}, clusters: {label: {sheetTitle, totalVolume, subClusters}}, breakdowns: {colonna:
{sheetTitle, totalVolume, values}} }, consumato da build_slides_pptx.py per il testo/i grafici
delle slide.

Uso:
    python scripts/build_sheet_xlsx.py \
        --volumes-csv runs/<slug>/volumes.csv \
        --clustered-csv runs/<slug>/clustering/clustered.csv \
        --xlsx-out runs/<slug>/output/report.xlsx \
        --charts-meta-out runs/<slug>/charts_meta.json \
        [--cluster-plan runs/<slug>/cluster_plan.json]
"""

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CLUSTERS_TAB = "clusters"
BRAND_SECCO_TAB = "brand-secco"
CLUSTER_OVERVIEW_TAB = "cluster-overview"

# Tab di esempio presenti nel template (prova di concetto usata per progettare il layout,
# vedi build_sheet.py originale) — rimossi a inizio run se ancora presenti nel template scaricato.
LEGACY_EXAMPLE_TABS = ["cluster_1", "Abbigliamento Donna"]

CLUSTERS_TAB_COLUMNS = [
    "Brand", "Country", "Keyword", "Position", "Search Volume", "Brand/Not Brand",
    "Cluster", "Sotto Cluster",
]
COL_SEARCH_VOLUME = CLUSTERS_TAB_COLUMNS.index("Search Volume")
COL_SOTTO_CLUSTER = CLUSTERS_TAB_COLUMNS.index("Sotto Cluster")

BREAKDOWN_TAB_PREFIX = "Extra "

CLUSTER_EFFECTIVE_COL = "Cluster Effettivo"
CLUSTER_OVERVIEW_COL = "Cluster Overview"

HEADER_FILL = PatternFill(start_color="1F4E77", end_color="1F4E77", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def attribute_columns_after_sotto_cluster(raw_columns) -> list[str]:
    """Colonne attributo dinamiche prodotte da claude-clustering-agent (Stagionalità, Genere,
    Brand correlati, ed eventuali nuove colonne future) — non una lista di nomi fissa nel
    codice, ma tutto cio' che nel CSV GREZZO (prima del riordino di reorder_clusters_columns)
    segue "Sotto Cluster". Stessa logica di build_sheet.py, invariata."""
    cols = list(raw_columns)
    if "Sotto Cluster" not in cols:
        return []
    return cols[cols.index("Sotto Cluster") + 1:]


def load_cluster_plan(path, clustered_df):
    if path:
        plan = json.loads(Path(path).read_text(encoding="utf-8"))
    else:
        plan = [{"label": str(c), "cluster": str(c), "genere": None}
                for c in clustered_df["Cluster"].dropna().unique()]

    labels = [entry["label"] for entry in plan]
    dupes = {label for label in labels if labels.count(label) > 1}
    if dupes:
        raise SystemExit(f"--cluster-plan ha etichette duplicate: {sorted(dupes)}")

    known_clusters = set(clustered_df["Cluster"].dropna().unique())
    unknown = sorted({entry["cluster"] for entry in plan if entry["cluster"] not in known_clusters})
    if unknown:
        raise SystemExit(f"--cluster-plan cita Cluster non presenti in --clustered-csv: {unknown}")

    return plan


def apply_cluster_plan(df: pd.DataFrame, plan) -> pd.DataFrame:
    df = df.copy()
    df[CLUSTER_EFFECTIVE_COL] = ""
    df[CLUSTER_OVERVIEW_COL] = ""
    genere_norm = df["Genere"].fillna("").astype(str) if "Genere" in df.columns else pd.Series([""] * len(df), index=df.index)

    for entry in plan:
        mask = df["Cluster"] == entry["cluster"]
        if entry.get("genere") is not None:
            wanted = {(g or "") for g in entry["genere"]}
            mask &= genere_norm.isin(wanted)
        df.loc[mask, CLUSTER_EFFECTIVE_COL] = entry["label"]
        df.loc[mask, CLUSTER_OVERVIEW_COL] = entry.get("overview_label", entry["label"])

    return df


def warn_unmatched_genere(df: pd.DataFrame, plan):
    split_clusters = {entry["cluster"] for entry in plan if entry.get("genere") is not None}
    for cluster in sorted(split_clusters):
        cluster_rows = df[df["Cluster"] == cluster]
        unmatched = cluster_rows[cluster_rows[CLUSTER_EFFECTIVE_COL] == ""]
        if len(unmatched):
            counts = unmatched["Genere"].fillna("(non specificato)").value_counts().to_dict()
            print(f"ATTENZIONE: {len(unmatched)} keyword del Cluster '{cluster}' non rientrano "
                  f"in nessuna etichetta del --cluster-plan e sono escluse da overview e tab "
                  f"({counts}) — verifica se e' voluto.")


def reorder_clusters_columns(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in CLUSTERS_TAB_COLUMNS if c not in df.columns]
    if missing:
        raise SystemExit(f"Colonne mancanti in --clustered-csv rispetto al template: {missing}")
    extra = [c for c in df.columns if c not in CLUSTERS_TAB_COLUMNS]
    return df[CLUSTERS_TAB_COLUMNS + extra]


def format_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def write_df(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.astype(object).where(pd.notna(df), "").itertuples(index=False):
        ws.append(list(row))
    format_header(ws, len(df.columns))


def write_brand_secco(ws, years, volumes):
    header = [f"Volume {y}" for y in years]
    ws.append(header)
    ws.append(volumes)
    ws.append([f"=(A2-B2)/B2", f"=(B2-C2)/C2"])
    for col in (1, 2):
        ws.cell(row=3, column=col).number_format = "0.00%"
    format_header(ws, len(header))


def aggregate(df: pd.DataFrame, group_col: str, value_label: str):
    """Aggregazione statica (sostituisce la PivotTable nativa): somma Search Volume per
    valore unico di group_col, ordinata per volume decrescente. Ritorna (records, total)."""
    grouped = (
        df.groupby(group_col)["Search Volume"]
        .sum()
        .sort_values(ascending=False)
    )
    records = []
    total = 0
    for key, vol in grouped.items():
        label = key if key not in (None, "") else "(non specificato)"
        vol = int(vol)
        records.append({value_label: label, "Volume Totale": vol})
        total += vol
    return records, total


def write_aggregation_tab(ws, records, value_label, chart_title):
    ws.append([value_label, "Volume Totale"])
    for rec in records:
        ws.append([rec[value_label], rec["Volume Totale"]])
    format_header(ws, 2)

    if not records:
        return

    chart = PieChart()
    chart.title = chart_title
    n = len(records)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=1 + n)
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.height = 10
    chart.width = 16
    ws.add_chart(chart, f"{get_column_letter(4)}1")


def breakdown_tab_title(column):
    return f"{BREAKDOWN_TAB_PREFIX}{column.replace('/', '-')}"[:31]


def main():
    parser = argparse.ArgumentParser(description="Genera l'xlsx completo del report (offline, nessuna chiamata Google)")
    parser.add_argument("--template-xlsx", required=True,
                         help="Percorso locale del Google Sheet template esportato in xlsx "
                              "(scaricato da Claude via mcp__claude_ai_Google_Drive__download_file_content)")
    parser.add_argument("--volumes-csv", required=True, help="CSV con colonne Anno,Search Volume (ultimi 3 anni, stesso periodo)")
    parser.add_argument("--clustered-csv", required=True, help="CSV finale di claude-clustering-agent")
    parser.add_argument("--xlsx-out", required=True, help="Percorso del report .xlsx generato")
    parser.add_argument("--charts-meta-out", required=True)
    parser.add_argument("--cluster-plan",
                         help='JSON [{"label", "cluster", "genere": [...]|null}, ...]. Se omesso: '
                              'un tab per ciascun valore unico di Cluster, nessuna esclusione ne\' split.')
    args = parser.parse_args()

    volumes_df = pd.read_csv(args.volumes_csv).sort_values("Anno", ascending=False)
    if len(volumes_df) != 3:
        raise SystemExit(f"--volumes-csv deve avere esattamente 3 righe (anno corrente + 2 precedenti), trovate {len(volumes_df)}")
    years = volumes_df["Anno"].astype(str).tolist()
    volumes = volumes_df["Search Volume"].astype(int).tolist()

    raw_clustered_df = pd.read_csv(args.clustered_csv)
    breakdown_columns = attribute_columns_after_sotto_cluster(raw_clustered_df.columns)
    clustered_df = reorder_clusters_columns(raw_clustered_df)
    plan = load_cluster_plan(args.cluster_plan, clustered_df)
    clustered_df = apply_cluster_plan(clustered_df, plan)
    warn_unmatched_genere(clustered_df, plan)

    wb = load_workbook(args.template_xlsx)

    # Rimuove i tab che questo script ricrea da zero (se già presenti nel template scaricato,
    # es. da un run precedente riusato come template) e gli eventuali tab di esempio residui —
    # preserva invece ogni altro tab del template (istruzioni, stile) non gestito qui.
    managed_tabs = {BRAND_SECCO_TAB, CLUSTERS_TAB, CLUSTER_OVERVIEW_TAB, *LEGACY_EXAMPLE_TABS}
    for name in list(wb.sheetnames):
        if name in managed_tabs or name.startswith(BREAKDOWN_TAB_PREFIX):
            wb.remove(wb[name])

    ws_brand_secco = wb.create_sheet(BRAND_SECCO_TAB)
    write_brand_secco(ws_brand_secco, years, volumes)

    ws_clusters = wb.create_sheet(CLUSTERS_TAB)
    write_df(ws_clusters, clustered_df)

    used_titles = {BRAND_SECCO_TAB, CLUSTERS_TAB, CLUSTER_OVERVIEW_TAB}
    overview_labels = list(dict.fromkeys(entry.get("overview_label", entry["label"]) for entry in plan))
    overview_df = clustered_df[clustered_df[CLUSTER_OVERVIEW_COL].isin(overview_labels)]
    overview_records, overview_total = aggregate(overview_df, CLUSTER_OVERVIEW_COL, "Cluster")

    ws_overview = wb.create_sheet(CLUSTER_OVERVIEW_TAB)
    write_aggregation_tab(ws_overview, overview_records, "Cluster", "Volume di ricerca per Cluster")

    charts_meta = {
        "overview": {"sheetTitle": CLUSTER_OVERVIEW_TAB, "totalVolume": overview_total, "subClusters": overview_records},
        "clusters": {},
    }

    for entry in plan:
        label = entry["label"]
        title = label[:31]
        suffix = 2
        while title in used_titles:
            title = f"{label[:28]} {suffix}"
            suffix += 1
        used_titles.add(title)

        cluster_df = clustered_df[clustered_df[CLUSTER_EFFECTIVE_COL] == label]
        records, total_volume = aggregate(cluster_df, "Sotto Cluster", "Sotto Cluster")

        ws_cluster = wb.create_sheet(title)
        write_aggregation_tab(ws_cluster, records, "Sotto Cluster", f"{label} — Ripartizione Sotto Cluster")

        charts_meta["clusters"][label] = {
            "sheetTitle": title,
            "totalVolume": total_volume,
            "subClusters": records,
        }

    charts_meta["breakdowns"] = {}
    for column in breakdown_columns:
        if clustered_df[column].fillna("").astype(str).str.strip().eq("").all():
            print(f"'{column}': colonna interamente vuota, nessun tab/grafico di breakdown creato.")
            continue

        title = breakdown_tab_title(column)
        used_titles.add(title)
        records, total_volume = aggregate(clustered_df, column, column)

        ws_breakdown = wb.create_sheet(title)
        write_aggregation_tab(ws_breakdown, records, column, f"Distribuzione per {column}")

        charts_meta["breakdowns"][column] = {
            "sheetTitle": title,
            "totalVolume": total_volume,
            "values": records,
        }

    xlsx_out = Path(args.xlsx_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_out)

    Path(args.charts_meta_out).write_text(json.dumps(charts_meta, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Fogli creati: {BRAND_SECCO_TAB}, {CLUSTERS_TAB}, {CLUSTER_OVERVIEW_TAB}, "
          f"+ {len(charts_meta['clusters'])} tab cluster + {len(charts_meta['breakdowns'])} tab breakdown extra")
    print(f"xlsx: {xlsx_out}")
    print(f"charts_meta: {args.charts_meta_out}")


if __name__ == "__main__":
    main()
